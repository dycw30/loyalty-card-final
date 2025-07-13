from flask import Flask, render_template, request, redirect, send_file, session, url_for
import sqlite3
from datetime import datetime
import os
import pandas as pd
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import qrcode

app = Flask(__name__)
app.secret_key = 'supersecretkey'
DB_NAME = "orders.db"
ADMIN_PASS = "adminpass2025"

def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            unique_id TEXT,
            drink_type TEXT,
            quantity INTEGER,
            tokens INTEGER,
            redeemed INTEGER,
            barista_name TEXT,
            date TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS baristas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password TEXT
        )
    """)
    conn.commit()
    c.execute("SELECT COUNT(*) FROM baristas")
    if c.fetchone()[0] == 0:
        c.executemany("INSERT INTO baristas (username, password) VALUES (?, ?)", [
            ('david', 'coffee123'),
            ('alice', 'latte456'),
            ('bob', 'mocha789')
        ])
        conn.commit()
    c.execute("PRAGMA table_info(orders)")
    columns = [col[1] for col in c.fetchall()]
    if 'barista_name' not in columns:
        c.execute("ALTER TABLE orders ADD COLUMN barista_name TEXT")
        conn.commit()
    conn.close()

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("SELECT * FROM baristas WHERE username=? AND password=?", (username, password))
        user = c.fetchone()
        conn.close()
        if user:
            session["logged_in"] = True
            session["barista"] = username
            return redirect(url_for("index"))
        else:
            return render_template("login.html", error="Invalid credentials")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/", methods=["GET", "POST"])
def index():
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    summary, top_customers, top_drinks = {}, [], []
    error = None
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    if request.method == "POST":
        unique_id = request.form["unique_id"]
        drink = request.form.get("drink") or None
        quantity = int(request.form["quantity"]) if request.form.get("quantity") else 0
        tokens = quantity // 9
        redeemed = int(request.form["redeem"]) if request.form.get("redeem") else 0
        barista_name = session.get("barista")
        date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        c.execute("""
            SELECT COALESCE(SUM(quantity),0)/9, COALESCE(SUM(redeemed),0)
            FROM orders WHERE unique_id=?
        """, (unique_id,))
        result = c.fetchone()
        total_tokens = int(result[0])
        total_redeemed = int(result[1])
        balance = total_tokens - total_redeemed

        if redeemed > balance:
            error = f"Cannot redeem {redeemed} tokens. Only {balance} available."
        else:
            c.execute("INSERT INTO orders (unique_id, drink_type, quantity, tokens, redeemed, barista_name, date) VALUES (?, ?, ?, ?, ?, ?, ?)",
                      (unique_id, drink, quantity, tokens, redeemed, barista_name, date))
            conn.commit()
            return redirect("/")

    c.execute("SELECT COUNT(DISTINCT unique_id), SUM(quantity), SUM(redeemed) FROM orders")
    result = c.fetchone()
    total_orders = result[1] or 0
    total_tokens = total_orders // 9
    summary = {
        "total_customers": result[0] or 0,
        "total_tokens": total_tokens,
        "total_redeemed": result[2] or 0,
        "balance_tokens": total_tokens - (result[2] or 0)
    }
    c.execute("""
        SELECT unique_id, SUM(quantity)/9 as tokens
        FROM orders
        GROUP BY unique_id
        ORDER BY tokens DESC LIMIT 5
    """)
    top_customers = [(row[0], int(row[1])) for row in c.fetchall()]
    c.execute("""
        SELECT drink_type, SUM(quantity)
        FROM orders
        WHERE drink_type IS NOT NULL
        GROUP BY drink_type
        ORDER BY SUM(quantity) DESC LIMIT 5
    """)
    top_drinks = c.fetchall()
    conn.close()
    return render_template("index.html", summary=summary, top_customers=top_customers, top_drinks=top_drinks, error=error)

@app.route("/export")
def export():
    if not session.get("logged_in"):
        return redirect(url_for("login"))
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    df = pd.read_sql_query("""
        SELECT unique_id,
               SUM(quantity) AS total_orders,
               SUM(quantity)/9 AS total_tokens,
               SUM(redeemed) AS total_redeemed,
               (SUM(quantity)/9) - SUM(redeemed) AS balance
        FROM orders
        GROUP BY unique_id
    """, conn)
    drink_df = pd.read_sql_query("""
        SELECT unique_id, drink_type, SUM(quantity) AS qty
        FROM orders
        WHERE drink_type IS NOT NULL
        GROUP BY unique_id, drink_type
    """, conn)
    if not drink_df.empty:
        drink_pivot = drink_df.pivot(index="unique_id", columns="drink_type", values="qty").fillna(0).astype(int)
    else:
        drink_pivot = pd.DataFrame()
    if not drink_pivot.empty:
        df = df.set_index("unique_id").join(drink_pivot).reset_index()
    conn.close()
    if df.empty:
        return "No data to export"
    wb = load_workbook("Bound CRM Test 3.xlsm", keep_vba=True)
    ws = wb["Sheet1"]
    headers = ["Unique ID", "Total Orders", "Tokens Earned", "Redeemed", "Balance"]
    drink_headers = list(drink_pivot.columns) if not drink_pivot.empty else []
    for idx, header in enumerate(headers + drink_headers, start=1):
        ws.cell(row=1, column=idx, value=header)
    start_row = 2
    for i, row in df.iterrows():
        ws.cell(row=start_row + i, column=1, value=row["unique_id"])
        ws.cell(row=start_row + i, column=2, value=int(row["total_orders"]))
        ws.cell(row=start_row + i, column=3, value=int(row["total_tokens"]))
        ws.cell(row=start_row + i, column=4, value=int(row["total_redeemed"]))
        ws.cell(row=start_row + i, column=5, value=int(row["balance"]))
        for j, drink in enumerate(drink_headers, start=6):
            ws.cell(row=start_row + i, column=j, value=int(row.get(drink, 0)))
    wb.save("Bound Cafe Aggregated Export.xlsm")
    return send_file("Bound Cafe Aggregated Export.xlsm", as_attachment=True)

@app.route("/pdf/<unique_id>")
def generate_pdf(unique_id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("""
        SELECT SUM(quantity), SUM(quantity)/9, SUM(redeemed), (SUM(quantity)/9)-SUM(redeemed)
        FROM orders WHERE unique_id=?
    """, (unique_id,))
    result = c.fetchone()
    total_orders = int(result[0] or 0)
    total_tokens = int(result[1] or 0)
    total_redeemed = int(result[2] or 0)
    balance = int(result[3] or 0)
    c.execute("""
        SELECT drink_type, SUM(quantity)
        FROM orders
        WHERE unique_id=?
        GROUP BY drink_type
    """, (unique_id,))
    drinks = c.fetchall()
    conn.close()
    filename = f"Loyalty_Statement_{unique_id}.pdf"
    pdf = canvas.Canvas(filename, pagesize=A4)
    width, height = A4
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(50, height - 50, f"Loyalty Statement for Customer: {unique_id}")
    pdf.setFont("Helvetica", 12)
    pdf.drawString(50, height - 80, f"Total Orders: {total_orders}")
    pdf.drawString(50, height - 100, f"Tokens Earned: {total_tokens}")
    pdf.drawString(50, height - 120, f"Tokens Redeemed: {total_redeemed}")
    pdf.drawString(50, height - 140, f"Balance Tokens: {balance}")
    y = height - 180
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(50, y, "Drink Breakdown:")
    y -= 20
    pdf.setFont("Helvetica", 12)
    for drink, qty in drinks:
        pdf.drawString(70, y, f"{drink}: {qty}")
        y -= 20
    pdf.save()
    return send_file(filename, as_attachment=True)

@app.route("/generate_qr/<unique_id>")
def generate_qr(unique_id):
    url = f"https://loyalty-card-app-v2.onrender.com/scan/{unique_id}"
    img = qrcode.make(url)
    filename = f"QR_{unique_id}.png"
    img.save(filename)
    return send_file(filename, as_attachment=True)

@app.route("/scan/<unique_id>")
def scan(unique_id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("""
        SELECT SUM(quantity), SUM(quantity)/9, SUM(redeemed), (SUM(quantity)/9)-SUM(redeemed)
        FROM orders WHERE unique_id=?
    """, (unique_id,))
    result = c.fetchone()
    total_orders = int(result[0] or 0)
    total_tokens = int(result[1] or 0)
    total_redeemed = int(result[2] or 0)
    balance = int(result[3] or 0)
    c.execute("""
        SELECT drink_type, SUM(quantity)
        FROM orders
        WHERE unique_id=?
        GROUP BY drink_type
    """, (unique_id,))
    drinks = c.fetchall()
    conn.close()
    return render_template("scan.html", unique_id=unique_id, total_orders=total_orders,
                           total_tokens=total_tokens, total_redeemed=total_redeemed,
                           balance=balance, drinks=drinks)

@app.route("/admin", methods=["GET", "POST"])
def admin():
    if request.method == "POST":
        master_password = request.form.get("master_password")
        if master_password != ADMIN_PASS:
            return render_template("admin.html", error="Invalid admin password", baristas=[])
        username = request.form.get("username")
        password = request.form.get("password")
        if username and password:
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            try:
                c.execute("INSERT INTO baristas (username, password) VALUES (?, ?)", (username, password))
                conn.commit()
            except sqlite3.IntegrityError:
                pass
            conn.close()
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT username FROM baristas")
    baristas = [row[0] for row in c.fetchall()]
    conn.close()
    return render_template("admin.html", baristas=baristas)

if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
